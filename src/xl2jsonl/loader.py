from __future__ import annotations

import csv
import logging
from pathlib import Path

import openpyxl
from python_calamine import CalamineWorkbook

from xl2jsonl.exceptions import LoaderError
from xl2jsonl.models import CellValue, SheetData

logger = logging.getLogger(__name__)

EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsb"}
CSV_EXTENSIONS = {".csv": ",", ".tsv": "\t"}
SUPPORTED_EXTENSIONS = EXCEL_EXTENSIONS | set(CSV_EXTENSIONS)


def load_workbook(
    path: Path,
    sheets: list[str | int] | None = None,
    engine: str = "auto",
) -> list[SheetData]:
    path = Path(path)
    if not path.exists():
        raise LoaderError(f"File not found: {path}")

    suffix = path.suffix.lower()

    if suffix in CSV_EXTENSIONS:
        if sheets is not None:
            logger.warning("Sheet selection ignored for %s files", suffix)
        return [_load_csv(path, delimiter=CSV_EXTENSIONS[suffix])]

    if suffix not in EXCEL_EXTENSIONS:
        raise LoaderError(
            f"Unsupported file format '{suffix}'. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    return _load_excel(path, sheets=sheets, engine=engine)


def _load_excel(
    path: Path,
    sheets: list[str | int] | None = None,
    engine: str = "auto",
) -> list[SheetData]:
    calamine_wb = CalamineWorkbook.from_path(str(path))
    all_sheet_names = calamine_wb.sheet_names
    target_sheets = _resolve_sheet_selection(all_sheet_names, sheets)

    # openpyxl only supports .xlsx
    supports_openpyxl = path.suffix.lower() == ".xlsx"

    openpyxl_wb = None
    if supports_openpyxl and engine in ("auto", "openpyxl"):
        openpyxl_wb = openpyxl.load_workbook(str(path), data_only=True)
    elif not supports_openpyxl and engine == "openpyxl":
        logger.warning(
            "openpyxl does not support %s files — using calamine instead",
            path.suffix,
        )

    results: list[SheetData] = []
    try:
        for idx, name in target_sheets:
            if engine == "calamine" or (not supports_openpyxl and engine != "auto"):
                results.append(_load_sheet_calamine(calamine_wb, name, idx))
            elif engine == "openpyxl" and openpyxl_wb is not None:
                results.append(_load_sheet_openpyxl(openpyxl_wb, name, idx))
            elif openpyxl_wb is not None:  # auto with openpyxl available
                ws = openpyxl_wb[name]
                merged_ranges = list(ws.merged_cells.ranges)
                if merged_ranges:
                    logger.info(
                        "Sheet '%s': %d merged region(s), using openpyxl",
                        name,
                        len(merged_ranges),
                    )
                    results.append(_load_sheet_openpyxl(openpyxl_wb, name, idx))
                else:
                    logger.debug("Sheet '%s': no merges, using calamine", name)
                    results.append(_load_sheet_calamine(calamine_wb, name, idx))
            else:  # auto without openpyxl (xls/xlsb)
                logger.debug(
                    "Sheet '%s': using calamine (no merge detection for %s)",
                    name,
                    path.suffix,
                )
                results.append(_load_sheet_calamine(calamine_wb, name, idx))
    finally:
        if openpyxl_wb:
            openpyxl_wb.close()

    return results


def _load_csv(path: Path, delimiter: str = ",") -> SheetData:
    """Load a CSV/TSV file as a single-sheet SheetData."""
    rows: list[list[CellValue]] = []

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            rows.append([_infer_csv_type(cell) for cell in row])

    return SheetData(name=path.stem, index=0, rows=rows)


def _infer_csv_type(value: str) -> CellValue:
    """Best-effort type inference for CSV cell values."""
    if not value:
        return None
    # Integer
    try:
        return int(value)
    except ValueError:
        pass
    # Float
    try:
        return float(value)
    except ValueError:
        pass
    # Boolean
    if value.lower() in ("true", "false"):
        return value.lower() == "true"
    return value


def _resolve_sheet_selection(
    all_names: list[str],
    sheets: list[str | int] | None,
) -> list[tuple[int, str]]:
    if sheets is None:
        return list(enumerate(all_names))

    result: list[tuple[int, str]] = []
    for s in sheets:
        if isinstance(s, int):
            if s < 0 or s >= len(all_names):
                raise LoaderError(
                    f"Sheet index {s} out of range (0-{len(all_names) - 1})"
                )
            result.append((s, all_names[s]))
        else:
            if s not in all_names:
                raise LoaderError(
                    f"Sheet '{s}' not found. Available: {all_names}"
                )
            result.append((all_names.index(s), s))
    return result


def _load_sheet_calamine(
    wb: CalamineWorkbook,
    sheet_name: str,
    sheet_idx: int,
) -> SheetData:
    data = wb.get_sheet_by_name(sheet_name).to_python()
    rows: list[list[CellValue]] = [list(row) for row in data]
    return SheetData(name=sheet_name, index=sheet_idx, rows=rows)


def _load_sheet_openpyxl(
    wb: openpyxl.Workbook,
    sheet_name: str,
    sheet_idx: int,
) -> SheetData:
    ws = wb[sheet_name]
    rows, had_merges = _resolve_merges(ws)
    return SheetData(
        name=sheet_name, index=sheet_idx, rows=rows, had_merged_cells=had_merges
    )


def _resolve_merges(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[list[list[CellValue]], bool]:
    merged_ranges = list(ws.merged_cells.ranges)
    had_merges = len(merged_ranges) > 0

    # Collect merge info before unmerging
    merge_fills: list[tuple[object, CellValue]] = []
    for mr in merged_ranges:
        value = ws.cell(mr.min_row, mr.min_col).value
        merge_fills.append((mr, value))

    # Unmerge all ranges
    for mr, _ in merge_fills:
        ws.unmerge_cells(str(mr))

    # Fill every cell in each former merged range
    for mr, value in merge_fills:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                ws.cell(row, col).value = value

    # Convert to 2D list
    rows: list[list[CellValue]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    return rows, had_merges
