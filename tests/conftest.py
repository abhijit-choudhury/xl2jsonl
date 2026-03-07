from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    """Simple 2-sheet workbook, no merges."""
    path = tmp_path / "simple.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "People"
    ws1.append(["Name", "Age", "City"])
    ws1.append(["Alice", 30, "London"])
    ws1.append(["Bob", 25, "Paris"])
    ws1.append(["Charlie", 35, "Berlin"])

    ws2 = wb.create_sheet("Scores")
    ws2.append(["Subject", "Score"])
    ws2.append(["Math", 95])
    ws2.append(["English", 88])

    wb.save(path)
    return path


@pytest.fixture
def merged_cells_xlsx(tmp_path: Path) -> Path:
    """Workbook with various merged cell patterns."""
    path = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row with a merged column header spanning B1:C1
    ws["A1"] = "ID"
    ws["B1"] = "Revenue"
    ws.merge_cells("B1:C1")  # "Revenue" spans B and C
    ws["D1"] = "Profit"

    # Data rows
    ws["A2"] = 1
    ws["B2"] = 100
    ws["C2"] = 200
    ws["D2"] = 50

    ws["A3"] = 2
    ws["B3"] = 150
    ws["C3"] = 250
    ws["D3"] = 75

    # Merged rows: A4:A5 merged (category spanning 2 rows)
    ws["A4"] = "Total"
    ws.merge_cells("A4:A5")
    ws["B4"] = 250
    ws["C4"] = 450
    ws["D4"] = 125
    ws["B5"] = 300
    ws["C5"] = 500
    ws["D5"] = 150

    wb.save(path)
    return path


@pytest.fixture
def empty_rows_xlsx(tmp_path: Path) -> Path:
    """Workbook with empty leading rows and gaps."""
    path = tmp_path / "empty_rows.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Empty rows 1-2, header at row 3
    ws["A3"] = "Product"
    ws["B3"] = "Price"
    ws["A4"] = "Widget"
    ws["B4"] = 9.99
    # Empty row 5
    ws["A6"] = "Gadget"
    ws["B6"] = 19.99

    wb.save(path)
    return path


@pytest.fixture
def mixed_types_xlsx(tmp_path: Path) -> Path:
    """Workbook with various data types."""
    import datetime

    path = tmp_path / "mixed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Name", "Active", "Score", "Date", "Notes"])
    ws.append(["Alice", True, 95.5, datetime.date(2024, 1, 15), "Line1\nLine2"])
    ws.append(["Bob", False, 88, datetime.datetime(2024, 6, 1, 14, 30), "  padded  "])

    wb.save(path)
    return path


@pytest.fixture
def merged_block_xlsx(tmp_path: Path) -> Path:
    """Workbook with a 2x2 merged block."""
    path = tmp_path / "merged_block.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["C1"] = "Header3"

    # 2x2 merged block at A2:B3
    ws["A2"] = "Block"
    ws.merge_cells("A2:B3")
    ws["C2"] = 10
    ws["C3"] = 20

    wb.save(path)
    return path
